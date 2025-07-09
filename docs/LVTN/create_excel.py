#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script tạo file Excel từ markdown database description
Yêu cầu: pip install openpyxl
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_markdown_tables():
    """Parse markdown file và trích xuất thông tin bảng"""
    
    with open('description.md', 'r', encoding='utf-8') as f:
        content = f.read()
    
    tables = []
    lines = content.split('\n')
    
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # Tìm tiêu đề bảng
        if line.startswith('## Bảng '):
            table_name = line.replace('## Bảng ', '').strip()
            
            # Tìm mô tả
            i += 1
            description = ""
            while i < len(lines) and not lines[i].strip().startswith('|'):
                if lines[i].strip().startswith('**Mô tả:**'):
                    description = lines[i].strip().replace('**Mô tả:**', '').strip()
                i += 1
            
            # Parse bảng markdown
            columns = []
            if i < len(lines) and '| Thuộc tính' in lines[i]:
                i += 1  # Skip header
                if i < len(lines) and ('|---' in lines[i] or '| :---' in lines[i]):
                    i += 1  # Skip separator
                
                # Đọc dữ liệu
                while i < len(lines):
                    data_line = lines[i].strip()
                    if not data_line or not data_line.startswith('|'):
                        break
                    
                    parts = [p.strip() for p in data_line.split('|')[1:-1]]
                    if len(parts) >= 6:
                        columns.append({
                            'attribute': parts[0],
                            'type': parts[1], 
                            'key': parts[2],
                            'unique': parts[3],
                            'mandatory': parts[4],
                            'description': parts[5]
                        })
                    i += 1
            
            tables.append({
                'name': table_name,
                'description': description,
                'columns': columns
            })
        else:
            i += 1
    
    return tables

def create_excel_file(tables):
    """Tạo file Excel với formatting đẹp"""
    
    wb = Workbook()
    
    # Xóa sheet mặc định
    wb.remove(wb.active)
    
    # Tạo sheet tổng quan
    overview_sheet = wb.create_sheet("Tổng quan")
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Data style
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Tạo sheet tổng quan
    overview_headers = ["STT", "Tên Bảng", "Mô tả", "Số cột"]
    for col, header in enumerate(overview_headers, 1):
        cell = overview_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Thêm dữ liệu tổng quan
    for row, table in enumerate(tables, 2):
        cell1 = overview_sheet.cell(row=row, column=1)
        cell1.value = row-1
        cell1.border = border

        cell2 = overview_sheet.cell(row=row, column=2)
        cell2.value = table['name']
        cell2.border = border

        cell3 = overview_sheet.cell(row=row, column=3)
        cell3.value = table['description']
        cell3.border = border

        cell4 = overview_sheet.cell(row=row, column=4)
        cell4.value = len(table['columns'])
        cell4.border = border
    
    # Auto-fit columns
    for col in range(1, 5):
        overview_sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # Tạo sheet cho từng bảng
    for table in tables:
        # Tạo tên sheet hợp lệ (Excel giới hạn 31 ký tự)
        sheet_name = table['name'][:31]
        ws = wb.create_sheet(sheet_name)
        
        # Tiêu đề bảng
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"Bảng: {table['name']}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Mô tả
        ws.merge_cells('A2:F2')
        desc_cell = ws['A2']
        desc_cell.value = f"Mô tả: {table['description']}"
        desc_cell.alignment = Alignment(horizontal="left", wrap_text=True)
        desc_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Headers
        headers = ["Thuộc tính", "Kiểu dữ liệu", "Khóa", "Duy nhất", "Bắt buộc", "Diễn giải"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Data rows
        for row, column in enumerate(table['columns'], 5):
            cell1 = ws.cell(row=row, column=1)
            cell1.value = column['attribute']
            cell1.border = border

            cell2 = ws.cell(row=row, column=2)
            cell2.value = column['type']
            cell2.border = border

            cell3 = ws.cell(row=row, column=3)
            cell3.value = column['key']
            cell3.border = border

            cell4 = ws.cell(row=row, column=4)
            cell4.value = column['unique']
            cell4.border = border

            cell5 = ws.cell(row=row, column=5)
            cell5.value = column['mandatory']
            cell5.border = border

            desc_cell = ws.cell(row=row, column=6)
            desc_cell.value = column['description']
            desc_cell.border = border
            desc_cell.alignment = data_alignment
        
        # Auto-fit columns
        column_widths = [25, 15, 8, 10, 10, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Set row height for description
        ws.row_dimensions[2].height = 30
    
    return wb

def main():
    """Hàm chính"""
    print("Đang parse file markdown...")
    tables = parse_markdown_tables()
    print(f"Đã tìm thấy {len(tables)} bảng")
    
    print("Đang tạo file Excel...")
    wb = create_excel_file(tables)
    
    # Lưu file
    filename = "csv_output/database_description.xlsx"
    wb.save(filename)
    print(f"Đã tạo file Excel: {filename}")
    
    # Tạo file CSV đơn giản cho từng bảng
    import csv
    import os
    
    for table in tables:
        csv_filename = f"csv_output/{table['name']}_simple.csv"
        with open(csv_filename, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow(['Thuộc tính', 'Kiểu dữ liệu', 'Khóa', 'Duy nhất', 'Bắt buộc', 'Diễn giải'])
            
            # Data
            for column in table['columns']:
                writer.writerow([
                    column['attribute'],
                    column['type'],
                    column['key'],
                    column['unique'],
                    column['mandatory'],
                    column['description']
                ])
    
    print("Hoàn thành! Các file đã được tạo:")
    print("- database_description.xlsx: File Excel với formatting đẹp")
    print("- [tên_bảng]_simple.csv: File CSV đơn giản cho từng bảng")

if __name__ == "__main__":
    main()
